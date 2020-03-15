import xlsxwriter
import openpyxl


def check_sheet_exists(sheet_name):
    try:
        f = open(sheet_name)
        exists = True
    except:
        exists, f = False, None
    finally:
        if f is not None:
            f.close()
    return exists


def createSheet(name, data):
    workbook = xlsxwriter.Workbook(name)
    worksheet = workbook.add_worksheet()
    cols = 0
    for i in data:
        worksheet.write(0, cols, i)
        cols += 1
    # row = 1
    # col = 1
    # for i in emailList:
    #     worksheet.write(row, col, i)
    #     row += 1
    # row = 1
    # for i in searchPhone:
    #     worksheet.write(row, col + 1, i)
    #     row += 1
    try:
        workbook.close()
        return True
    except:
        print("Couldn't create xlsx file")


def appendSheet(sheetName, data):
    wb = openpyxl.load_workbook(sheetName)
    sheetList = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetList[0])
    maxRow = sheet.max_row
    temp = maxRow
    dataKey = list(data.keys())
    for i in range(0, len(dataKey)):
        cell = sheet.cell(row=(temp + 1), column=i+1)
        cell.value = data[dataKey[i]]
    wb.save(sheetName)